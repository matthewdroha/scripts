#!/usr/bin/env python3
"""find_collisions — detect RTL module definition collisions from a VCS elaboration XML dump.

See scripts/rtl/find_collisions.spec.md for the full specification.

Deviations from the spec text, verified against a real config_diagnostics.xml
(14.6 GB uncompressed / 16,269,991 <Instance> elements -- see
/memories/repo/find_collisions.md):
  - find_collisions.csv is FILTERED to real collisions only: a module name
    counts as colliding when it is bound to more than one distinct
    (library, configrule, source) combination. A full, unfiltered inventory
    would be ~78K+ rows on real data, almost all non-colliding.
  - Every <Instance> has exactly one <DefinitionDetails> (Module/Library/
    SourceInfo) and one <ConfigRule><Rule>; only <InstanceDetails> vs.
    <TopDetails> (the design top) differs, and neither is needed here.
  - The file has no XML declaration and pretty-prints tags with a trailing
    space before '>' (e.g. `<Module >x</Module >`) -- both are ordinary,
    well-formed XML and need no special handling.
  - find_collisions.minimized.csv rolls the same colliding modules up by
    (module, source) instead of the full 4-tuple, reporting how many distinct
    libraries/configrules were bound to that module+source pair. Both reports
    are derived from a single XML parse (build_definition_counts()) -- the
    16M-instance file is only streamed once per run.
  - find_collisions.minimized.realpath.csv further collapses find_collisions.
    minimized.csv's `source` field via os.path.realpath, so two textually
    different `../../..`-style relative paths that resolve to the same real
    file no longer look like distinct definitions. A module that no longer has
    more than one distinct (library, configrule, canonical-source) combination
    after this collapse is dropped -- the same ">1" collision rule, just
    reapplied post-canonicalization. Only already-colliding modules (per the
    raw counts) are canonicalized (canonicalizing can only ever MERGE distinct
    combinations, never split them, so a non-colliding module can't become
    colliding here) -- this keeps the (NFS-touching) realpath() calls bounded.
  - Both minimized reports get one extra 1/0 column per DISTINCT configrule
    value found ANYWHERE in the whole XML (not just among colliding rows) --
    see §3.2/§3.3/§6 Q8. This is the tool's one external dependency:
    `openpyxl` (pip install --user openpyxl) is also required to write
    find_collisions.xlsx (§3.4), breaking the previous stdlib-only guarantee.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import os
import re
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterator

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DefKey = tuple[str, str, str, str]  # module, library, configrule, source


@dataclass(frozen=True)
class Config:
    xml_path: Path
    output_root: Path

    @property
    def report_csv(self) -> Path:
        return self.output_root / "find_collisions.csv"

    @property
    def minimized_csv(self) -> Path:
        return self.output_root / "find_collisions.minimized.csv"

    @property
    def xlsx_path(self) -> Path:
        return self.output_root / "find_collisions.xlsx"

    @property
    def minimized_realpath_csv(self) -> Path:
        return self.output_root / "find_collisions.minimized.realpath.csv"


# --------------------------------------------------------------------------- #
# XML streaming
# --------------------------------------------------------------------------- #
def _is_gzip(path: Path) -> bool:
    with path.open("rb") as fh:
        return fh.read(2) == b"\x1f\x8b"


def open_xml(path: Path):
    """Open path for reading, transparently handling gzip regardless of extension."""
    if _is_gzip(path):
        return gzip.open(path, "rb")
    return path.open("rb")


def _text(elem: ET.Element, tag: str) -> str:
    child = elem.find(tag)
    return child.text.strip() if child is not None and child.text else ""


def iter_definitions(
    xml_path: Path,
    *,
    verbose: bool = False,
    progress_interval: int = 1_000_000,
    log: Callable[[str], None] = print,
) -> Iterator[DefKey]:
    """Stream-parse the elaboration XML, yielding a DefKey per <Instance>.

    Clears each processed <Instance> from the in-memory tree as it goes (the
    standard iterparse "clear the root" trick) so memory stays bounded
    regardless of file size.
    """
    count = 0
    with open_xml(xml_path) as fh:
        context = ET.iterparse(fh, events=("start", "end"))
        _, root = next(context)
        for event, elem in context:
            if event != "end" or elem.tag != "Instance":
                continue
            definition = elem.find("DefinitionDetails")
            config_rule = elem.find("ConfigRule")
            if definition is not None and config_rule is not None:
                yield (
                    _text(definition, "Module"),
                    _text(definition, "Library"),
                    _text(config_rule, "Rule"),
                    _text(definition, "SourceInfo"),
                )
            count += 1
            root.clear()
            if verbose and count % progress_interval == 0:
                log(f"-I- processed {count:,} instances...")
    if verbose:
        log(f"-I- processed {count:,} instances total.")


# --------------------------------------------------------------------------- #
# Collision derivation
# --------------------------------------------------------------------------- #
_FIELDS = ["module", "library", "configrule", "source", "instance_count"]
_MINIMIZED_FIELDS = ["module", "source", "library_count", "config_rule_count", "instance_count"]


def build_definition_counts(
    xml_path: Path, *, verbose: bool = False, log: Callable[[str], None] = print,
) -> Counter[DefKey]:
    """Parse the XML once into instance counts per (module, library, configrule, source).

    Both collect_collision_rows() and collect_minimized_rows() derive from this
    single Counter so a run only streams the (potentially 14+ GB) XML once.
    """
    return Counter(iter_definitions(xml_path, verbose=verbose, log=log))


def colliding_modules(counts: Counter[DefKey]) -> set[str]:
    """Module names bound to more than one distinct (library, configrule, source).

    This is find_collisions.csv's (§3.1) collision rule -- any field differing
    counts, even a library/configrule-only difference on an identical source.
    """
    by_module: dict[str, set[tuple[str, str, str]]] = {}
    for module, library, rule, source in counts:
        by_module.setdefault(module, set()).add((library, rule, source))
    return {m for m, defs in by_module.items() if len(defs) > 1}


def modules_with_multiple_sources(counts: Counter[DefKey]) -> set[str]:
    """Module names bound to more than one DISTINCT source, ignoring library/configrule.

    This is the (stricter) qualification rule for the minimized reports (§3.2/
    §3.3): a module bound to a single source under multiple library/configrule
    labels is NOT a real collision here, even though it is one for
    colliding_modules()/find_collisions.csv. Always a subset of
    colliding_modules() (a source difference alone already makes the
    (library, configrule, source) tuples distinct).
    """
    by_module: dict[str, set[str]] = {}
    for module, _library, _rule, source in counts:
        by_module.setdefault(module, set()).add(source)
    return {m for m, sources in by_module.items() if len(sources) > 1}


def collect_collision_rows(counts: Counter[DefKey]) -> list[dict]:
    """Build the filtered (collisions-only) row list for find_collisions.csv."""
    colliding = colliding_modules(counts)
    rows = [
        {
            "module": module,
            "library": library,
            "configrule": rule,
            "source": source,
            "instance_count": n,
        }
        for (module, library, rule, source), n in counts.items()
        if module in colliding
    ]
    rows.sort(key=lambda r: (r["module"], -r["instance_count"]))
    return rows


def all_configrule_values(counts: Counter[DefKey]) -> list[str]:
    """All distinct configrule values found anywhere in the whole-XML counts, sorted.

    `counts` (from build_definition_counts()) covers every <Instance> in the
    XML, not just colliding ones, so this is every configrule value that
    exists in the design -- used as the minimized reports' extra 1/0 columns
    (§3.2/§3.3/§6 Q8), even for a configrule value that never appears among
    the colliding/qualifying rows themselves.
    """
    return sorted({rule for (_module, _library, rule, _source) in counts})


def collect_minimized_rows(counts: Counter[DefKey], all_configrules: list[str] | None = None) -> list[dict]:
    """Build the (module, source)-keyed row list for find_collisions.minimized*.csv.

    Scoped to modules with more than one DISTINCT SOURCE (see
    modules_with_multiple_sources()) -- a library/configrule-only difference on
    an identical source does not qualify here. For each qualifying module+source
    pair, reports how many distinct libraries/configrules were bound to it and
    the total instance count across those bindings.

    If `all_configrules` is given (see all_configrule_values()), each row also
    gets one 1/0 column per configrule value in that list -- 1 iff that
    configrule was one of the ones bound to this module+source pair (so the
    sum of those columns equals the row's config_rule_count).
    """
    qualifying = modules_with_multiple_sources(counts)
    agg: dict[tuple[str, str], dict] = {}
    for (module, library, rule, source), n in counts.items():
        if module not in qualifying:
            continue
        entry = agg.setdefault((module, source), {"libraries": set(), "configrules": set(), "instance_count": 0})
        entry["libraries"].add(library)
        entry["configrules"].add(rule)
        entry["instance_count"] += n

    rows = []
    for (module, source), entry in agg.items():
        row = {
            "module": module,
            "source": source,
            "library_count": len(entry["libraries"]),
            "config_rule_count": len(entry["configrules"]),
            "instance_count": entry["instance_count"],
        }
        if all_configrules is not None:
            for rule in all_configrules:
                row[rule] = 1 if rule in entry["configrules"] else 0
        rows.append(row)
    rows.sort(key=lambda r: (r["module"], -r["instance_count"]))
    return rows


_SOURCE_RE = re.compile(r'^"(.*)",(\d+)$')


def canonicalize_source(source: str) -> str:
    """Resolve a `"<path>",<linenum>` SourceInfo string's path via os.path.realpath.

    Collapses textually-different-but-equivalent relative-path traversals (e.g.
    differing `../../..` segments, or symlinked directories) that resolve to the
    same real file. Works even if the path no longer exists on disk (realpath
    falls back to lexical normalization for the non-existent portion, same as
    its usual behavior). Strings that don't match the expected `"path",linenum`
    shape are returned unchanged.
    """
    m = _SOURCE_RE.match(source)
    if not m:
        return source
    path, linenum = m.group(1), m.group(2)
    return f'"{os.path.realpath(path)}",{linenum}'


def build_realpath_counts(counts: Counter[DefKey], modules: set[str] | None = None) -> Counter[DefKey]:
    """Re-key counts with each tuple's `source` canonicalized via realpath.

    Tuples that collapse to the same (module, library, configrule,
    canonical_source) key after canonicalization have their instance counts
    summed. `modules` restricts canonicalization to a subset (default: all) --
    pass the already-colliding module set to avoid realpath() calls (which may
    touch NFS) for the vast majority of non-colliding tuples; canonicalization
    can only ever MERGE distinct combinations, never split them, so a module
    that wasn't colliding before can't become colliding after.
    """
    cache: dict[str, str] = {}
    realpath_counts: Counter[DefKey] = Counter()
    for (module, library, rule, source), n in counts.items():
        if modules is not None and module not in modules:
            continue
        canonical = cache.get(source)
        if canonical is None:
            canonical = canonicalize_source(source)
            cache[source] = canonical
        realpath_counts[(module, library, rule, canonical)] += n
    return realpath_counts


def write_report(rows: list[dict], path: Path, fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel_workbook(path: Path, sheets: list[tuple[str, list[dict], list[str]]]) -> None:
    """Write one .xlsx with one sheet per (sheet_name, rows, fields) tuple.

    Each sheet's header+data range is wrapped in an Excel Table object (§3.4)
    so the data can be sorted/filtered natively in Excel. Header is always
    written, even for zero rows.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows, fields in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        last_col = get_column_letter(len(fields))
        last_row = len(rows) + 1
        table = Table(displayName=f"{sheet_name}_table", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False,
        )
        ws.add_table(table)
    wb.save(path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def preflight(cfg: Config) -> list[str]:
    errors = []
    if not cfg.xml_path.is_file():
        errors.append(f"--xml file not found: {cfg.xml_path}")
    return errors


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="find_collisions.py",
        description="Detect RTL module definition collisions from a VCS elaboration XML dump.",
    )
    parser.add_argument("--xml", required=True, type=Path, help="VCS elaboration XML dump (may be gzipped).")
    parser.add_argument("--dry-run", action="store_true", help="Print the plan; write nothing.")
    parser.add_argument(
        "--force", action="store_true",
        help="No-op: the report is always regenerated/overwritten each run. Kept for CLI parity.",
    )
    parser.add_argument("--verbose", action="store_true", help="Log progress and the file written.")
    return parser


def resolve_config(args: argparse.Namespace) -> Config:
    return Config(xml_path=args.xml.resolve(), output_root=Path.cwd())


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    cfg = resolve_config(args)

    errors = preflight(cfg)
    if errors:
        for e in errors:
            print(f"-E- {e}", file=sys.stderr)
        return 2

    if args.dry_run:
        print("Would write:")
        print(f"  {cfg.report_csv}")
        print(f"  {cfg.minimized_csv}")
        print(f"  {cfg.minimized_realpath_csv}")
        print(f"  {cfg.xlsx_path}")
        return 0

    counts = build_definition_counts(cfg.xml_path, verbose=args.verbose)
    configrules = all_configrule_values(counts)
    minimized_fields = _MINIMIZED_FIELDS + configrules
    rows = collect_collision_rows(counts)
    minimized_rows = collect_minimized_rows(counts, configrules)
    realpath_counts = build_realpath_counts(counts, modules=colliding_modules(counts))
    minimized_realpath_rows = collect_minimized_rows(realpath_counts, configrules)
    write_report(rows, cfg.report_csv, _FIELDS)
    write_report(minimized_rows, cfg.minimized_csv, minimized_fields)
    write_report(minimized_realpath_rows, cfg.minimized_realpath_csv, minimized_fields)
    write_excel_workbook(cfg.xlsx_path, [
        ("collisions", rows, _FIELDS),
        ("minimized", minimized_rows, minimized_fields),
        ("minimized_realpath", minimized_realpath_rows, minimized_fields),
    ])

    if args.verbose:
        print(f"wrote {cfg.report_csv} ({len(rows)} row(s))")
        print(f"wrote {cfg.minimized_csv} ({len(minimized_rows)} row(s))")
        print(f"wrote {cfg.minimized_realpath_csv} ({len(minimized_realpath_rows)} row(s))")
        print(f"wrote {cfg.xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
