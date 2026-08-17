#!/usr/bin/env python3
"""Unit tests for find_collisions.py.

Fixtures build small synthetic elaboration XML documents (both plain and
gzipped) rather than depending on any real config_diagnostics.xml. See
scripts/rtl/find_collisions.spec.md §7 for the test plan this covers.
"""

from __future__ import annotations

import gzip
import tempfile
import unittest
from pathlib import Path

import find_collisions as fc

TOP_INSTANCE = """\
<Instance >
<TopDetails >
<Hierarchy >top</Hierarchy >
<SourceInfo >"/src/top.sv",1</SourceInfo >
</TopDetails >
<DefinitionDetails >
<Module >top</Module >
<Library >TOP_LIB</Library >
<SourceInfo >"/src/top.sv",1</SourceInfo >
</DefinitionDetails >
<ConfigRule >
<Rule >Top Module</Rule >
</ConfigRule >
</Instance >
"""


def _instance(hierarchy, module, library, source, rule):
    return f"""\
<Instance >
<InstanceDetails >
<Hierarchy >{hierarchy}</Hierarchy >
<SourceInfo >{source}</SourceInfo >
</InstanceDetails >
<DefinitionDetails >
<Module >{module}</Module >
<Library >{library}</Library >
<SourceInfo >{source}</SourceInfo >
</DefinitionDetails >
<ConfigRule >
<Rule >{rule}</Rule >
</ConfigRule >
</Instance >
"""


def build_xml(instances: str) -> str:
    return f"<InstanceList >\n{TOP_INSTANCE}{instances}</InstanceList >\n"


def _counts(instances: str):
    """Write a fixture XML (top instance + given instances) and return its Counter."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "x.xml"
        path.write_text(build_xml(instances), encoding="utf-8")
        return fc.build_definition_counts(path)


class TestOpenXml(unittest.TestCase):
    def test_plain_and_gzip_both_readable(self):
        with tempfile.TemporaryDirectory() as tmp:
            plain = Path(tmp) / "a.xml"
            plain.write_text(build_xml(""), encoding="utf-8")
            with fc.open_xml(plain) as fh:
                self.assertIn(b"InstanceList", fh.read())

            gz = Path(tmp) / "a.xml.gz"
            with gzip.open(gz, "wb") as out:
                out.write(build_xml("").encode("utf-8"))
            with fc.open_xml(gz) as fh:
                self.assertIn(b"InstanceList", fh.read())

    def test_gzip_detected_without_gz_suffix(self):
        with tempfile.TemporaryDirectory() as tmp:
            gz_no_suffix = Path(tmp) / "a.dump"
            with gzip.open(gz_no_suffix, "wb") as out:
                out.write(build_xml("").encode("utf-8"))
            with fc.open_xml(gz_no_suffix) as fh:
                self.assertIn(b"InstanceList", fh.read())


class TestIterDefinitions(unittest.TestCase):
    def test_yields_one_tuple_per_instance_including_top(self):
        xml = build_xml(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "default library search order")
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "x.xml"
            path.write_text(xml, encoding="utf-8")
            defs = list(fc.iter_definitions(path))
        self.assertEqual(
            defs,
            [
                ("top", "TOP_LIB", "Top Module", '"/src/top.sv",1'),
                ("mod_a", "LIB_A", "default library search order", '"/src/a.sv",5'),
            ],
        )

    def test_verbose_logs_progress_and_total(self):
        xml = build_xml(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "default library search order")
        )
        logged = []
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "x.xml"
            path.write_text(xml, encoding="utf-8")
            list(fc.iter_definitions(path, verbose=True, progress_interval=1, log=logged.append))
        self.assertTrue(any("processed 1 instances..." in m for m in logged))
        self.assertTrue(any("processed 2 instances total." in m for m in logged))


class TestCollectCollisionRows(unittest.TestCase):
    def test_module_with_single_definition_is_excluded(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
        )
        rows = fc.collect_collision_rows(counts)
        # "top" (unique) and "mod_a" (single definition, 2 instances) both non-colliding.
        self.assertEqual(rows, [])

    def test_module_bound_to_two_libraries_is_a_collision(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.c", "mod_a", "LIB_B", '"/src/a.sv",5', "parent cell's library")
        )
        rows = fc.collect_collision_rows(counts)
        self.assertEqual(
            rows,
            [
                {
                    "module": "mod_a", "library": "LIB_B", "configrule": "parent cell's library",
                    "source": '"/src/a.sv",5', "instance_count": 2,
                },
                {
                    "module": "mod_a", "library": "LIB_A", "configrule": "parent cell's library",
                    "source": '"/src/a.sv",5', "instance_count": 1,
                },
            ],
        )

    def test_same_library_and_source_but_different_configrule_is_a_collision(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_A", '"/src/a.sv",5', "default library search order")
        )
        rows = fc.collect_collision_rows(counts)
        self.assertEqual(len(rows), 2)
        self.assertEqual({r["configrule"] for r in rows}, {"parent cell's library", "default library search order"})

    def test_sort_is_module_then_descending_instance_count(self):
        counts = _counts(
            _instance("top.a", "mod_z", "LIB_A", '"/src/z.sv",1', "r1")
            + _instance("top.b", "mod_z", "LIB_B", '"/src/z.sv",2', "r1")
            + _instance("top.c", "mod_z", "LIB_B", '"/src/z.sv",2', "r1")
            + _instance("top.d", "mod_a", "LIB_A", '"/src/a.sv",1', "r1")
            + _instance("top.e", "mod_a", "LIB_B", '"/src/a.sv",2', "r1")
        )
        rows = fc.collect_collision_rows(counts)
        modules_in_order = [r["module"] for r in rows]
        self.assertEqual(modules_in_order, ["mod_a", "mod_a", "mod_z", "mod_z"])
        # Within mod_z: LIB_B (count 2) before LIB_A (count 1).
        self.assertEqual(
            [r["library"] for r in rows if r["module"] == "mod_z"], ["LIB_B", "LIB_A"],
        )


class TestModulesWithMultipleSources(unittest.TestCase):
    def test_same_source_different_library_is_not_multi_source(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "r1")
        )
        self.assertEqual(fc.modules_with_multiple_sources(counts), set())

    def test_different_source_is_multi_source(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
            + _instance("top.b", "mod_a", "LIB_A", '"/src/b.sv",5', "r1")
        )
        self.assertEqual(fc.modules_with_multiple_sources(counts), {"mod_a"})

    def test_is_always_a_subset_of_colliding_modules(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "r1")
            + _instance("top.c", "mod_b", "LIB_A", '"/src/b.sv",5', "r1")
            + _instance("top.d", "mod_b", "LIB_A", '"/src/c.sv",5', "r1")
        )
        self.assertTrue(fc.modules_with_multiple_sources(counts) <= fc.colliding_modules(counts))


class TestCollectMinimizedRows(unittest.TestCase):
    def test_module_with_single_definition_is_excluded(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
            + _instance("top.b", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
        )
        self.assertEqual(fc.collect_minimized_rows(counts), [])

    def test_same_source_different_library_does_not_qualify(self):
        """Reproduces the reported false positive: 2 different libraries bound to
        the exact same source is a collision for find_collisions.csv, but must
        NOT appear in the minimized reports (only >1 DISTINCT SOURCE qualifies).
        """
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "parent cell's library")
        )
        self.assertEqual(fc.collect_collision_rows(counts) != [], True)  # DOES collide for the raw report
        self.assertEqual(fc.collect_minimized_rows(counts), [])  # does NOT qualify here

    def test_groups_by_module_and_source_counting_distinct_libraries_and_rules(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.c", "mod_a", "LIB_B", '"/src/a.sv",5', "default library search order")
            + _instance("top.d", "mod_a", "LIB_A", '"/src/b.sv",9', "parent cell's library")
        )
        rows = fc.collect_minimized_rows(counts)
        self.assertEqual(
            rows,
            [
                {
                    "module": "mod_a", "source": '"/src/a.sv",5',
                    "library_count": 2, "config_rule_count": 2, "instance_count": 3,
                },
                {
                    "module": "mod_a", "source": '"/src/b.sv",9',
                    "library_count": 1, "config_rule_count": 1, "instance_count": 1,
                },
            ],
        )

    def test_sort_is_module_then_descending_instance_count(self):
        counts = _counts(
            _instance("top.a", "mod_z", "LIB_A", '"/src/z.sv",1', "r1")
            + _instance("top.b", "mod_z", "LIB_B", '"/src/z.sv",1', "r1")
            + _instance("top.c", "mod_z", "LIB_B", '"/src/z2.sv",1', "r1")
            + _instance("top.d", "mod_a", "LIB_A", '"/src/a.sv",1', "r1")
            + _instance("top.e", "mod_a", "LIB_B", '"/src/a.sv",1', "r1")
            + _instance("top.f", "mod_a", "LIB_B", '"/src/a2.sv",1', "r1")
        )
        rows = fc.collect_minimized_rows(counts)
        self.assertEqual([r["module"] for r in rows], ["mod_a", "mod_a", "mod_z", "mod_z"])
        # Within mod_z: z.sv (instance_count 2) sorts before z2.sv (instance_count 1).
        self.assertEqual(
            [r["source"] for r in rows if r["module"] == "mod_z"], ['"/src/z.sv",1', '"/src/z2.sv",1'],
        )

    def test_configrule_columns_are_1_0_and_sum_to_config_rule_count(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "parent cell's library")
            + _instance("top.c", "mod_a", "LIB_B", '"/src/b.sv",9', "default library search order")
        )
        all_configrules = fc.all_configrule_values(counts)
        # includes the top-instance's own "Top Module" rule too, whole-XML universe.
        self.assertEqual(all_configrules, ["Top Module", "default library search order", "parent cell's library"])
        rows = fc.collect_minimized_rows(counts, all_configrules)
        by_source = {r["source"]: r for r in rows}
        row_a = by_source['"/src/a.sv",5']
        self.assertEqual(row_a["parent cell's library"], 1)
        self.assertEqual(row_a["default library search order"], 0)
        self.assertEqual(row_a["Top Module"], 0)
        self.assertEqual(
            sum(row_a[r] for r in all_configrules), row_a["config_rule_count"],
        )
        row_b = by_source['"/src/b.sv",9']
        self.assertEqual(row_b["default library search order"], 1)
        self.assertEqual(
            sum(row_b[r] for r in all_configrules), row_b["config_rule_count"],
        )

    def test_configrule_columns_omitted_when_all_configrules_not_given(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
            + _instance("top.b", "mod_a", "LIB_B", '"/src/b.sv",9', "r1")
        )
        rows = fc.collect_minimized_rows(counts)
        self.assertNotIn("r1", rows[0])


class TestAllConfigruleValues(unittest.TestCase):
    def test_covers_whole_xml_even_non_colliding_modules(self):
        counts = _counts(
            _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "only rule")
        )
        # "only rule" appears on a non-colliding module, and "Top Module" is the
        # fixture's always-present top instance -- both must still show up.
        self.assertEqual(fc.all_configrule_values(counts), ["Top Module", "only rule"])


class TestWriteReport(unittest.TestCase):
    def test_write_report_header_always_present(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "find_collisions.csv"
            fc.write_report([], path, fc._FIELDS)
            content = path.read_text(encoding="utf-8")
        self.assertEqual(content.strip(), "module,library,configrule,source,instance_count")

    def test_write_report_rows(self):
        rows = [
            {
                "module": "mod_a", "library": "LIB_A", "configrule": "r1",
                "source": '"/src/a.sv",5', "instance_count": 3,
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "find_collisions.csv"
            fc.write_report(rows, path, fc._FIELDS)
            lines = path.read_text(encoding="utf-8").splitlines()
        self.assertEqual(lines[0], "module,library,configrule,source,instance_count")
        self.assertIn('mod_a,LIB_A,r1,"""/src/a.sv"",5",3', lines[1])

    def test_write_minimized_report_header_and_rows(self):
        rows = [
            {
                "module": "mod_a", "source": '"/src/a.sv",5',
                "library_count": 2, "config_rule_count": 1, "instance_count": 3,
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "find_collisions.minimized.csv"
            fc.write_report(rows, path, fc._MINIMIZED_FIELDS)
            lines = path.read_text(encoding="utf-8").splitlines()
        self.assertEqual(lines[0], "module,source,library_count,config_rule_count,instance_count")
        self.assertIn('mod_a,"""/src/a.sv"",5",2,1,3', lines[1])


class TestWriteExcelWorkbook(unittest.TestCase):
    def test_writes_one_table_per_sheet_including_zero_row_sheet(self):
        import openpyxl

        sheets = [
            ("collisions", [{"module": "mod_a", "instance_count": 3}], ["module", "instance_count"]),
            ("minimized", [], ["module", "source"]),  # zero rows -- header-only table must still work
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "find_collisions.xlsx"
            fc.write_excel_workbook(path, sheets)
            wb = openpyxl.load_workbook(path)
            self.assertEqual(wb.sheetnames, ["collisions", "minimized"])

            ws1 = wb["collisions"]
            self.assertEqual([c.value for c in ws1[1]], ["module", "instance_count"])
            self.assertEqual([c.value for c in ws1[2]], ["mod_a", 3])
            self.assertEqual(list(ws1.tables.keys()), ["collisions_table"])
            self.assertEqual(ws1.tables["collisions_table"].ref, "A1:B2")

            ws2 = wb["minimized"]
            self.assertEqual([c.value for c in ws2[1]], ["module", "source"])
            self.assertEqual(list(ws2.tables.keys()), ["minimized_table"])
            self.assertEqual(ws2.tables["minimized_table"].ref, "A1:B1")


class TestCanonicalizeSource(unittest.TestCase):
    def test_collapses_relative_traversal(self):
        source = '"/nfs/site/x/subip/sip/ubr_d2d_cbb_bot_array_wrapper/../../../subip/sip/arf044_dfx_wrapper/src/rtl/arf044_map.sv",5'
        self.assertEqual(fc.canonicalize_source(source), '"/nfs/site/x/subip/sip/arf044_dfx_wrapper/src/rtl/arf044_map.sv",5')

    def test_two_different_traversals_resolve_to_same_canonical_source(self):
        a = '"/nfs/site/x/subip/sip/ubr_d2d_cbb_bot_array_wrapper/../../../subip/sip/arf044_dfx_wrapper/src/rtl/arf044_map.sv",5'
        b = '"/nfs/site/x/subip/sip/ubr_d2d_cbb_top_array_wrapper/../../../subip/sip/arf044_dfx_wrapper/src/rtl/arf044_map.sv",5'
        self.assertEqual(fc.canonicalize_source(a), fc.canonicalize_source(b))

    def test_different_linenum_stays_distinct(self):
        a = '"/nfs/site/x/subip/sip/wrap_a/../../../subip/sip/mod/src/rtl/mod.sv",5'
        b = '"/nfs/site/x/subip/sip/wrap_b/../../../subip/sip/mod/src/rtl/mod.sv",9'
        self.assertNotEqual(fc.canonicalize_source(a), fc.canonicalize_source(b))

    def test_unparsable_string_returned_unchanged(self):
        self.assertEqual(fc.canonicalize_source("not-quoted-path"), "not-quoted-path")


class TestBuildRealpathCounts(unittest.TestCase):
    def test_merges_tuples_that_canonicalize_to_the_same_key(self):
        a = '"/nfs/x/subip/sip/wrap_bot/../../../subip/sip/mod/src/rtl/mod.sv",5'
        b = '"/nfs/x/subip/sip/wrap_top/../../../subip/sip/mod/src/rtl/mod.sv",5'
        counts = fc.Counter({
            ("mod_a", "LIB_A", "r1", a): 3,
            ("mod_a", "LIB_A", "r1", b): 4,
        })
        realpath_counts = fc.build_realpath_counts(counts)
        self.assertEqual(len(realpath_counts), 1)
        self.assertEqual(next(iter(realpath_counts.values())), 7)

    def test_modules_filter_skips_other_modules(self):
        counts = fc.Counter({
            ("mod_a", "LIB_A", "r1", '"/a.sv",1'): 1,
            ("mod_b", "LIB_A", "r1", '"/b.sv",1'): 1,
        })
        realpath_counts = fc.build_realpath_counts(counts, modules={"mod_a"})
        self.assertEqual(list(realpath_counts.keys()), [("mod_a", "LIB_A", "r1", '"/a.sv",1')])


class TestPreflightAndCli(unittest.TestCase):
    def test_preflight_missing_xml(self):
        cfg = fc.Config(xml_path=Path("/no/such/file.xml"), output_root=Path("/tmp"))
        errors = fc.preflight(cfg)
        self.assertEqual(len(errors), 1)
        self.assertIn("--xml file not found", errors[0])

    def test_preflight_existing_xml_passes(self):
        with tempfile.TemporaryDirectory() as tmp:
            xml_path = Path(tmp) / "x.xml"
            xml_path.write_text(build_xml(""), encoding="utf-8")
            cfg = fc.Config(xml_path=xml_path, output_root=Path(tmp))
            self.assertEqual(fc.preflight(cfg), [])

    def test_main_missing_xml_exits_2(self):
        rc = fc.main(["--xml", "/no/such/file.xml"])
        self.assertEqual(rc, 2)

    def test_main_dry_run_writes_nothing(self):
        with tempfile.TemporaryDirectory() as tmp:
            xml_path = Path(tmp) / "x.xml"
            xml_path.write_text(
                build_xml(
                    _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
                    + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "r1")
                ),
                encoding="utf-8",
            )
            outdir = Path(tmp) / "out"
            outdir.mkdir()
            import os
            old_cwd = os.getcwd()
            os.chdir(outdir)
            try:
                rc = fc.main(["--xml", str(xml_path), "--dry-run"])
            finally:
                os.chdir(old_cwd)
            self.assertEqual(rc, 0)
            self.assertFalse((outdir / "find_collisions.csv").exists())
            self.assertFalse((outdir / "find_collisions.minimized.csv").exists())
            self.assertFalse((outdir / "find_collisions.minimized.realpath.csv").exists())
            self.assertFalse((outdir / "find_collisions.xlsx").exists())

    def test_main_writes_report_in_cwd(self):
        with tempfile.TemporaryDirectory() as tmp:
            xml_path = Path(tmp) / "x.xml"
            xml_path.write_text(
                build_xml(
                    _instance("top.a", "mod_a", "LIB_A", '"/src/a.sv",5', "r1")
                    + _instance("top.b", "mod_a", "LIB_B", '"/src/a.sv",5', "r1")
                    + _instance("top.c", "mod_a", "LIB_A", '"/src/b.sv",9', "r1")
                ),
                encoding="utf-8",
            )
            outdir = Path(tmp) / "out"
            outdir.mkdir()
            import os
            old_cwd = os.getcwd()
            os.chdir(outdir)
            try:
                rc = fc.main(["--xml", str(xml_path)])
            finally:
                os.chdir(old_cwd)
            self.assertEqual(rc, 0)
            report = outdir / "find_collisions.csv"
            self.assertTrue(report.exists())
            lines = report.read_text(encoding="utf-8").splitlines()
            self.assertEqual(len(lines), 4)  # header + 3 colliding rows (distinct library/source combos)

            # Both fixture instances use configrule "r1"; the top instance (always
            # present via build_xml()) uses "Top Module" -- so those are the two
            # whole-XML configrule columns appended to the minimized reports.
            minimized = outdir / "find_collisions.minimized.csv"
            self.assertTrue(minimized.exists())
            min_lines = minimized.read_text(encoding="utf-8").splitlines()
            self.assertEqual(
                min_lines[0],
                "module,source,library_count,config_rule_count,instance_count,Top Module,r1",
            )
            self.assertEqual(len(min_lines), 3)  # header + 2 distinct-source rows
            self.assertIn('mod_a,"""/src/a.sv"",5",2,1,2,0,1', min_lines[1])

            # Same 2 distinct sources -> realpath collapsing changes nothing here
            # (no relative traversal to collapse) -- see the dedicated
            # relative-path-collapse test below for the actual bug scenario.
            minimized_realpath = outdir / "find_collisions.minimized.realpath.csv"
            self.assertTrue(minimized_realpath.exists())
            realpath_lines = minimized_realpath.read_text(encoding="utf-8").splitlines()
            self.assertEqual(
                realpath_lines[0],
                "module,source,library_count,config_rule_count,instance_count,Top Module,r1",
            )
            self.assertEqual(len(realpath_lines), 3)  # header + 2 distinct-source rows
            self.assertIn('mod_a,"""/src/a.sv"",5",2,1,2,0,1', realpath_lines[1])

            import openpyxl

            xlsx_path = outdir / "find_collisions.xlsx"
            self.assertTrue(xlsx_path.exists())
            wb = openpyxl.load_workbook(xlsx_path)
            self.assertEqual(wb.sheetnames, ["collisions", "minimized", "minimized_realpath"])

    def test_relative_path_collapse_drops_module_from_realpath_report(self):
        """Reproduces the reported bug: a module whose only real difference is a
        `../../..`-style relative path spelling shows up in find_collisions.csv
        and find_collisions.minimized.csv (raw text differs) but must NOT appear
        in find_collisions.minimized.realpath.csv once the paths are collapsed to
        the same real file (>1 rule no longer applies).
        """
        source_a = (
            '"/nfs/site/x/subip/sip/ubr_d2d_cbb_bot_array_wrapper/../../../subip/sip/'
            'arf044_dfx_wrapper/src/rtl/arf044_map.sv",5'
        )
        source_b = (
            '"/nfs/site/x/subip/sip/ubr_d2d_cbb_top_array_wrapper/../../../subip/sip/'
            'arf044_dfx_wrapper/src/rtl/arf044_map.sv",5'
        )
        with tempfile.TemporaryDirectory() as tmp:
            xml_path = Path(tmp) / "x.xml"
            xml_path.write_text(
                build_xml(
                    _instance("top.a", "arf044_clk", "LIB_A", source_a, "r1")
                    + _instance("top.b", "arf044_clk", "LIB_A", source_b, "r1")
                ),
                encoding="utf-8",
            )
            outdir = Path(tmp) / "out"
            outdir.mkdir()
            import os
            old_cwd = os.getcwd()
            os.chdir(outdir)
            try:
                rc = fc.main(["--xml", str(xml_path)])
            finally:
                os.chdir(old_cwd)
            self.assertEqual(rc, 0)

            report = (outdir / "find_collisions.csv").read_text(encoding="utf-8")
            self.assertIn("arf044_clk", report)
            minimized = (outdir / "find_collisions.minimized.csv").read_text(encoding="utf-8")
            self.assertIn("arf044_clk", minimized)

            realpath_report = (outdir / "find_collisions.minimized.realpath.csv").read_text(encoding="utf-8")
            self.assertNotIn("arf044_clk", realpath_report)
            self.assertEqual(
                realpath_report.strip(),
                "module,source,library_count,config_rule_count,instance_count,Top Module,r1",
            )


if __name__ == "__main__":
    unittest.main()
